#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_db.py - vygeneruje data/customers.json z "data odberatele.xlsx" a "data prodeje.xlsx".

Pouziti:
    python3 scripts/build_db.py \
        --odberatele "/cesta/data odberatele.xlsx" \
        --prodeje    "/cesta/data prodeje.xlsx" \
        --out        "data/customers.json" \
        --overrides  "data/overrides.json"

Skript NIKDY neprepisuje data/tasks.json, data/completed.json ani data/overrides.json.
Rucni zmeny obchodniho zastupce v overrides.json maji prednost pred automatickym prirazenim.
"""
import argparse, json, os, re, sys, unicodedata
from datetime import datetime, timedelta, timezone

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from psc_okres import psc_to_okres  # noqa: E402

import openpyxl  # noqa: E402

# ---------------------------------------------------------------- konfigurace
# kraje, ktere patri AHUY (vse ostatni -> NAM)
AHUY_KRAJE = {"Ústecký", "Liberecký", "Královéhradecký",
              "Hlavní město Praha", "Středočeský"}

NEW_CUSTOMER_DAYS = 90     # "khach moi" = vytvoren do 90 dni a bez objednavky
RECENT_ORDER_DAYS = 120    # "4 mesice"
TIER_MID = 50000           # hranice mezi nepotencialnim a beznym
TIER_HIGH = 300000         # hranice pro velkeho zakaznika

# skupina -> (poradi priority, interval navstevy ve dnech nebo None, text priority)
GROUPS = {
    "khach lon":        (1, 21,  "Návštěva 1x / 21 dní"),
    "khach tiem nang":  (2, 30,  "Návštěva 1x / 30 dní"),
    "khach moi":        (3, 90,  "Kontakt / návštěva do 3 měsíců od registrace"),
    "khach thuong":     (4, 60,  "Návštěva 1x / 60 dní"),
    "khach tham khao":  (5, None, "Kontakt kdykoli / návštěva při cestě"),
    "khach khong tiem nang": (6, None, "Kontakt kdykoli / návštěva při cestě"),
    "khach ngung mua":  (7, None, "Neodebírá – bez termínu návštěvy"),
}
# Skupiny, ktere nikdy nedostanou termin navstevy (interval je None).
# "khach ngung mua" = khach khong lay hang nua - nastavuje jen NAM rucne na webu.
# Ktere ukoly se pocitaji jako kontakt se zakaznikem (= "ngay lien lac cuoi")
# a resetuji termin navstevy. Jen "pozn" = ghi chú; reklamace (rekl) se nepocita.
KONTAKT_TYPY = {"pozn"}

GROUP_LABEL = {
    "khach lon": "Khách lớn",
    "khach tiem nang": "Khách tiềm năng",
    "khach moi": "Khách mới",
    "khach thuong": "Khách thường",
    "khach tham khao": "Khách tham khảo",
    "khach khong tiem nang": "Khách không tiềm năng",
    "khach ngung mua": "Khách không lấy hàng nữa",
}


# ------------------------------------------------------------------- pomocne
def norm_ico(v):
    if v is None:
        return ""
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    s = re.sub(r"\D", "", s)
    return s.zfill(8) if s else ""


def parse_dt(v):
    if isinstance(v, datetime):
        return v
    if not v:
        return None
    for f in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(str(v).strip()[:19], f)
        except ValueError:
            pass
    return None


def parse_iso(v):
    """ISO 8601 z webu ('2026-08-12T08:55:39.036Z') -> naivni local-like datetime."""
    if not v:
        return None
    s = str(v).strip().replace("Z", "+00:00")
    try:
        dt = datetime.fromisoformat(s)
    except ValueError:
        return None
    if dt.tzinfo is not None:
        dt = dt.astimezone().replace(tzinfo=None)
    return dt


def d(v):
    return v.strftime("%Y-%m-%d") if isinstance(v, datetime) else None


def extract_psc(addr):
    m = re.search(r"(\d{3})\s?(\d{2})\s", addr or "")
    return (m.group(1) + m.group(2)) if m else ""


def extract_mesto(addr):
    m = re.search(r"\d{3}\s?\d{2}\s+([^,]+)", addr or "")
    if not m:
        return ""
    return re.sub(r"\s*\(.*", "", m.group(1)).strip()


def split_contacts(*vals):
    """Rozdeli 'El. adresa 1/2' na telefon a email (poradi ve zdroji nemusi sedet)."""
    tel, mail = [], []
    for v in vals:
        s = (str(v) if v is not None else "").strip()
        if not s:
            continue
        for part in re.split(r"[;,]\s*", s):
            part = part.strip()
            if not part:
                continue
            if "@" in part:
                mail.append(part)
            elif re.search(r"\d{6,}", part):
                tel.append(part)
            else:
                mail.append(part)
    return ", ".join(dict.fromkeys(tel)), ", ".join(dict.fromkeys(mail))


def strip_diac(s):
    return "".join(c for c in unicodedata.normalize("NFD", s or "")
                   if unicodedata.category(c) != "Mn")


# --------------------------------------------------------------- nacteni dat
def read_odberatele(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    hdr = [str(c).strip() if c else "" for c in next(rows)]
    idx = {h: i for i, h in enumerate(hdr)}

    def g(r, name, default=""):
        i = idx.get(name)
        return r[i] if i is not None and i < len(r) and r[i] is not None else default

    out = []
    for r in rows:
        if not any(r):
            continue
        ico = norm_ico(g(r, "IČ"))
        firma = str(g(r, "Firma")).strip()
        if not firma:
            continue
        dodaci = str(g(r, "Dodací adresa")).strip()
        koresp = str(g(r, "Adresa (Korespondenční adresa)")).strip()
        addr = dodaci or koresp                      # priorita: dodaci adresa
        psc = extract_psc(addr) or extract_psc(koresp)
        okres, kraj = psc_to_okres(psc)
        tel, mail = split_contacts(g(r, "El. adresa 1"), g(r, "El. adresa 2"))
        created = parse_dt(g(r, "Vytvořeno", None))
        cislo = g(r, "Číslo", None)
        if not ico and not addr:
            continue                                  # prazdny interni zaznam
        out.append({
            "id": ico or ("c%s" % cislo),             # stabilni klic zaznamu
            "cislo": cislo,
            "ico": ico,
            "firma": firma,
            "cenova_skupina": str(g(r, "Cen. sk.")).strip(),
            "dodaci_nazev": str(g(r, "Dodací adresa - název")).strip(),
            "adresa": addr,
            "adresa_koresp": koresp,
            "psc": psc,
            "mesto": extract_mesto(addr) or extract_mesto(koresp),
            "okres": okres,
            "kraj": kraj,
            "telefon": tel,
            "email": mail,
            "pracovnik": str(g(r, "Pracovník")).strip(),
            "vytvoreno": d(created),
        })
    return out


def read_prodeje(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    hdr = [str(c).strip() if c else "" for c in next(rows)]
    idx = {h: i for i, h in enumerate(hdr)}
    i_ico = idx.get("Odběratel")
    i_net = idx.get("Netto M")          # sloupec F = netto castka
    i_dat = idx.get("Vytvořeno")
    i_dok = idx.get("Doklad")

    agg = {}
    for r in rows:
        if not any(r):
            continue
        ico = norm_ico(r[i_ico]) if i_ico is not None else ""
        if not ico:
            continue
        netto = r[i_net] if i_net is not None and isinstance(r[i_net], (int, float)) else 0
        dt = parse_dt(r[i_dat]) if i_dat is not None else None
        a = agg.setdefault(ico, {"pocet": 0, "obrat": 0.0, "prvni": None,
                                 "posledni": None, "posledni_doklad": ""})
        a["pocet"] += 1
        a["obrat"] += float(netto or 0)
        if dt:
            if a["prvni"] is None or dt < a["prvni"]:
                a["prvni"] = dt
            if a["posledni"] is None or dt > a["posledni"]:
                a["posledni"] = dt
                a["posledni_doklad"] = str(r[i_dok] or "") if i_dok is not None else ""
    return agg


# ----------------------------------------------------------- klasifikace
def classify(cust, today):
    pocet = cust["pocet_objednavek"]
    obrat = cust["obrat_celkem"]
    posl = parse_dt(cust["posledni_nakup"])
    vytv = parse_dt(cust["vytvoreno"])
    recent = bool(posl and (today - posl).days <= RECENT_ORDER_DAYS)

    if pocet == 0:
        if vytv and (today - vytv).days <= NEW_CUSTOMER_DAYS:
            return "khach moi"
        return "khach tham khao"
    if obrat >= TIER_HIGH:
        return "khach lon" if recent else "khach thuong"
    if obrat >= TIER_MID:
        return "khach tiem nang" if recent else "khach thuong"
    return "khach khong tiem nang"


def skupina_override(cust, ovr):
    """Rucne nastavena skupina (dropdown na webu, jen NAM).

    ovr = hodnota z overrides.json["skupina"][id]; bud retezec s klicem skupiny,
    nebo {"g": klic, "at": ISO datum zmeny, "by": jmeno}.
    Vraci (klic_skupiny | None, zdroj). Rucni volba PLATI VZDY, krome pripadu,
    kdy zakaznik po datu zmeny udelal novou objednavku - pak se vraci
    k automatickemu zarazeni (napr. "prestal odebirat" a zase zacal nakupovat).
    """
    if not ovr:
        return None, None
    if isinstance(ovr, str):
        g, at, by = ovr, None, ""
    else:
        g, at, by = ovr.get("g"), parse_iso(ovr.get("at")), ovr.get("by", "")
    if g not in GROUPS:
        return None, None
    posl = parse_dt(cust.get("posledni_nakup"))
    if at and posl and posl > at:
        return None, None                      # nova objednavka -> zpet na automat
    return g, {"at": d(at), "by": by}


def task_kontakt_datum(t):
    """Datum kontaktu jedne poznamky.

    Prednost ma rucne zadane datum ("kontakt" - policko na webu u ghi chú),
    jinak se bere datum zapisu / posledni upravy (u dokoncene poznamky done_at).
    """
    man = parse_dt(t.get("kontakt"))
    if man:
        return man, True
    dt = parse_iso(t.get("done_at")) or parse_iso(t.get("updated_at")) \
        or parse_iso(t.get("created_at"))
    return dt, False


def read_kontakty(*paths):
    """tasks.json + completed.json -> {cust_id: {...} } - posledni kontakt.

    Pocitaji se jen poznamky (KONTAKT_TYPY). Bere se poznamka s nejnovejsim
    datem kontaktu (rucne zadane datum ma prednost pred datem zapisu).
    """
    out = {}
    for path in paths:
        if not path or not os.path.exists(path):
            continue
        with open(path, encoding="utf-8") as f:
            tasks = json.load(f).get("tasks", [])
        for t in tasks:
            if t.get("type") not in KONTAKT_TYPY:
                continue
            cid = str(t.get("cust_id") or "")
            dt, rucne = task_kontakt_datum(t)
            if not cid or not dt:
                continue
            cur = out.get(cid)
            if cur is None or dt > cur["datum"]:
                out[cid] = {"datum": dt, "rucne": rucne,
                            "by": t.get("done_by") or t.get("updated_by")
                                  or t.get("created_by") or "",
                            "text": (t.get("text") or "")[:200],
                            "hotovo": bool(t.get("done_at"))}
    return out


def posledni_kontakt(cust, kont):
    """(datum, zdroj) posledniho kontaktu.

    1) datum z poznamky (rucne zadane, jinak datum zapisu poznamky)
    2) zakaznik bez poznamky -> datum posledniho nakupu
    3) bez nakupu (novy zakaznik) -> datum zalozeni
    """
    if kont:
        return kont["datum"], ("pozn_rucne" if kont["rucne"] else "pozn")
    posl = parse_dt(cust.get("posledni_nakup"))
    if posl:
        return posl, "nakup"
    vytv = parse_dt(cust.get("vytvoreno"))
    if vytv:
        return vytv, "vytvoreno"
    return None, None


def deadline(cust, today):
    """Datum, do kdy je treba kontakt, + kolik dni je po termínu.

    Zaklad = "ngay lien lac cuoi" (cust["posledni_kontakt"]), tedy datum
    z posledni poznamky; u zakaznika bez poznamky datum posledniho nakupu.
    """
    interval = GROUPS[cust["skupina"]][1]
    if interval is None:
        return None, 0
    base = parse_dt(cust.get("posledni_kontakt"))
    if not base:
        return None, 0
    due = base + timedelta(days=interval)
    return due.strftime("%Y-%m-%d"), max(0, (today - due).days)


# ---------------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--odberatele", required=True)
    ap.add_argument("--prodeje", required=True)
    ap.add_argument("--out", default="data/customers.json")
    ap.add_argument("--overrides", default="data/overrides.json")
    ap.add_argument("--previous", default=None,
                    help="predchozi customers.json - zachova prirazeni OZ")
    ap.add_argument("--completed", default=None,
                    help="data/completed.json - dokoncene poznamky = kontakt, resetuji termin")
    ap.add_argument("--tasks", default=None,
                    help="data/tasks.json - otevrene poznamky = kontakt, resetuji termin")
    a = ap.parse_args()

    today = datetime.now()
    overrides, skup_ovr = {}, {}
    if os.path.exists(a.overrides):
        with open(a.overrides, encoding="utf-8") as f:
            _ov = json.load(f)
        overrides = _ov.get("obchodni_zastupce", {})
        skup_ovr = _ov.get("skupina", {})            # rucne nastavena skupina (NAM)

    # predchozi verze databaze - drzi prirazeni obchodniho zastupce
    previous = {}
    if a.previous and os.path.exists(a.previous):
        with open(a.previous, encoding="utf-8") as f:
            previous = {c["id"]: c for c in json.load(f).get("customers", [])}

    odb = read_odberatele(a.odberatele)
    sales = read_prodeje(a.prodeje)
    kontakty = read_kontakty(a.tasks, a.completed)

    customers, warn_psc, warn_oz = [], [], []
    for o in odb:
        s = sales.get(o["ico"], {})
        c = dict(o)
        c["pocet_objednavek"] = s.get("pocet", 0)
        c["obrat_celkem"] = round(s.get("obrat", 0.0), 2)
        c["prvni_nakup"] = d(s.get("prvni"))
        c["posledni_nakup"] = d(s.get("posledni"))
        c["posledni_doklad"] = s.get("posledni_doklad", "")
        c["dni_od_nakupu"] = ((today - s["posledni"]).days
                              if s.get("posledni") else None)

        # ngay lien lac cuoi = datum posledni poznamky (rucne zadane ma prednost),
        # u zakaznika bez poznamky = datum posledniho nakupu
        kon = kontakty.get(c["id"])
        kdat, kzdroj = posledni_kontakt(c, kon)
        c["posledni_kontakt"] = d(kdat)
        c["kontakt_zdroj"] = kzdroj                  # pozn_rucne | pozn | nakup | vytvoreno
        c["kontakt_by"] = kon["by"] if kon else ""
        c["kontakt_text"] = kon["text"] if kon else ""
        c["dni_od_kontaktu"] = (today - kdat).days if kdat else None
        # zpetna kompatibilita se starsim webem
        c["posledni_navsteva"] = d(kon["datum"]) if kon else None
        c["navsteva_by"] = c["kontakt_by"]
        c["navsteva_text"] = c["kontakt_text"]
        c["dni_od_navstevy"] = ((today - kon["datum"]).days if kon else None)

        # --- skupina -----------------------------------------------------
        # Automaticky vypocet, ktery muze NAM prebit dropdownem na webu.
        # Rucni volba se automaticky nemeni (viz skupina_override).
        c["skupina_auto"] = classify(c, today)
        g_man, g_info = skupina_override(c, skup_ovr.get(c["id"]))
        c["skupina"] = g_man or c["skupina_auto"]
        c["skupina_rucne"] = bool(g_man)
        c["skupina_zmena"] = (g_info or {}).get("at")
        c["skupina_by"] = (g_info or {}).get("by", "")
        c["skupina_label"] = GROUP_LABEL[c["skupina"]]
        c["priorita"] = GROUPS[c["skupina"]][0]
        c["priorita_text"] = GROUPS[c["skupina"]][2]
        c["termin_do"], c["po_terminu_dni"] = deadline(c, today)

        # --- obchodni zastupce ------------------------------------------
        # Pravidlo podle kraje se pouzije POUZE u zakaznika, ktery jeste
        # v databazi nebyl. U znameho zakaznika se drzi drivejsi prirazeni,
        # i kdyby se mu zmenila adresa. Menit smi jen NAM (overrides.json).
        c["oz_kraj"] = "AHUY" if c["kraj"] in AHUY_KRAJE else "NAM"
        prev = previous.get(c["id"])
        c["oz_auto"] = prev.get("oz_auto", c["oz_kraj"]) if prev else c["oz_kraj"]
        c["obchodni_zastupce"] = overrides.get(c["id"], c["oz_auto"])
        c["oz_rucne"] = c["id"] in overrides
        c["oz_novy"] = prev is None
        if prev and c["oz_kraj"] != c["oz_auto"]:
            warn_oz.append((c["firma"], c["okres"], c["oz_auto"], c["oz_kraj"]))

        # pole pro fulltext vyhledavani (bez diakritiky)
        c["_search"] = strip_diac(" ".join([
            c["firma"], c["ico"], c["telefon"], c["email"], c["adresa"],
            c["dodaci_nazev"], c["okres"], c["kraj"], c["mesto"],
            c["skupina_label"], c["obchodni_zastupce"],
            c["posledni_nakup"] or "",
        ])).lower()

        if c["okres"] == "Neznámý":
            warn_psc.append((c["firma"], c["psc"], c["adresa"]))
        customers.append(c)

    customers.sort(key=lambda x: (x["priorita"], -(x["po_terminu_dni"] or 0),
                                  -x["obrat_celkem"]))

    payload = {
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "pocet_zakazniku": len(customers),
        "pravidla": {
            "novy_zakaznik_dni": NEW_CUSTOMER_DAYS,
            "aktivni_nakup_dni": RECENT_ORDER_DAYS,
            "hranice_stredni": TIER_MID,
            "hranice_velky": TIER_HIGH,
            "ahuy_kraje": sorted(AHUY_KRAJE),
        },
        "customers": customers,
    }
    os.makedirs(os.path.dirname(a.out) or ".", exist_ok=True)
    with open(a.out, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=1)

    # -- souhrn do konzole
    from collections import Counter
    print("Zapsano:", a.out, "-", len(customers), "zakazniku")
    for k, v in sorted(Counter(c["skupina_label"] for c in customers).items()):
        print("  %-24s %d" % (k, v))
    print("  OZ:", dict(Counter(c["obchodni_zastupce"] for c in customers)))
    rucne = [c for c in customers if c.get("skupina_rucne")]
    if rucne:
        print("  Skupina nastavena rucne (NAM na webu): %d" % len(rucne))
        for c in rucne[:30]:
            print("    %-38s -> %-26s (auto: %s)"
                  % (c["firma"][:38], c["skupina_label"], GROUP_LABEL[c["skupina_auto"]]))
    novi = [c for c in customers if c.get("oz_novy")]
    if previous:
        print("  OZ prirazen automaticky jen novym zakaznikum: %d" % len(novi))
    if warn_psc:
        print("!! Nerozpoznane PSC (%d):" % len(warn_psc))
        for w in warn_psc[:30]:
            print("   ", w)
    if warn_oz:
        print("i  Zakaznici, kde by pravidlo podle kraje dalo jineho OZ (NEMENENO,")
        print("   zmenit muze jen NAM na webu) - %d:" % len(warn_oz))
        for f, ok, drzi, kraj in warn_oz[:30]:
            print("    %-38s %-20s drzi=%-5s kraj by dal=%s" % (f[:38], ok, drzi, kraj))


if __name__ == "__main__":
    main()
